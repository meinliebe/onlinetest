from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, HttpResponseRedirect
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from django.contrib import messages
from django.contrib.auth.models import User
from django.views.decorators.csrf import csrf_protect
from django.db import transaction, DatabaseError, Error, IntegrityError
from django.db.models import Q
from django.core.exceptions import ObjectDoesNotExist, EmptyResultSet
from django.forms import inlineformset_factory
from django.conf import settings
from soal.models import Soal, Mata_kuliah, Jawaban, Kategori_soal

from soal.forms import SoalForm, JawabanForm, UploadPesertaForm, KategoriSoalForm, PesertaForm
from ujian.models import Peserta
from openpyxl import load_workbook
import json
# Create your views here.
def get_mata_kuliah(request):
	data = {}
	mata_kuliah = Mata_kuliah.objects.all()
	return mata_kuliah

@csrf_protect
@login_required(login_url = settings.LOGIN_URL)
def dashboard_admin(request):
	data = {}
	data['app_title'] = 'ADMIN CBT'

	return render(request, 'soal/dashboard_admin.html', data)

@csrf_protect
@login_required(login_url = settings.LOGIN_URL)
def list_soal(request):
	data = {}
	data['all'] = False
	data['app_title'] = 'ADMIN | DAFTAR SOAL'
	if request.GET:
		param_mata_kuliah = request.GET.get('mata_kuliah')
		param_kata_kunci = request.GET.get('kata_kunci')
		soal_obj = Soal.objects.filter(Q(text_soal__icontains = param_kata_kunci) )		
		data['all'] = True
		data['keyword'] = param_kata_kunci
	else:
		soal_obj = Soal.objects.all().order_by('-id')[:20]
	data['soal'] = soal_obj
	data['mata_kuliah'] = get_mata_kuliah(request)
	return render(request, 'soal/list_soal.html', data)

@csrf_protect
@login_required(login_url = settings.LOGIN_URL)
def view_soal(request, id_soal):
	data = {}
	get_soal = get_object_or_404(Soal, pk=id_soal)	
	data['app_title'] = 'ADMIN | VIEW SOAL'
	data['soal'] = get_soal
	return render(request, 'soal/view_soal.html', data)

@csrf_protect
@transaction.atomic
@login_required(login_url = settings.LOGIN_URL)
def ubah_soal(request, id_soal):
	data = {}
	data['app_title'] = 'ADMIN | UBAH SOAL'
	success = True

	get_soal = get_object_or_404(Soal, pk=id_soal)	
	get_jawaban = Jawaban.objects.filter(soal = get_soal)
	jawaban_benar = get_jawaban.get(jawaban_benar=True)
	count_jawaban = get_jawaban.count()
	JawabanFormSet = inlineformset_factory(Soal, Jawaban, fields=('text_jawaban', 'gambar','jawaban_benar', ), extra=(5-count_jawaban), can_delete=False)

	if request.POST:
		f = SoalForm(request.POST, request.FILES, instance=get_soal)
		if f.is_valid():
			soal_form = f.save(commit=False)
			soal_form.user_update = request.user.username
			try:
				with transaction.atomic():
					f.save()
					if soal_form.tipe_soal == 1:
						j = JawabanFormSet(request.POST, request.FILES, instance=soal_form)	
						j.save()
					else:
						jawaban = ['A', 'B', 'C', 'D', 'E']
						jawab = request.POST.get('jawaban')
						jawaban_lama = Jawaban.objects.filter(soal=get_soal)
						if jawaban_lama:
							for j_l in jawaban_lama:
								j_l.delete()
						for j in jawaban:
							if jawab == j.lower():
								benar = True
							else:
								benar = False
							Jawaban.objects.update_or_create(soal=get_soal, text_jawaban=j, defaults={'jawaban_benar': benar})
					messages.add_message(request, messages.SUCCESS, '<strong>Sukses! </strong> Data soal berhasil diubah')
					return redirect('view_soal', id_soal)
			except DatabaseError:
				messages.add_message(request, messages.WARNING, '<strong>Error! </strong> Terjadi masalah saat meyimpan data, silakan hub IT')					
				success = False

	data['soal_form'] = SoalForm(instance=get_soal)	
	data['jawaban_form'] = JawabanFormSet(instance=get_soal)
	data['jawaban_benar'] = jawaban_benar.text_jawaban
	data['soal_id'] = get_soal.id
	data['tipe_soal'] = get_soal.tipe_soal
	return render(request, 'soal/ubah_soal.html', data)

@csrf_protect
@transaction.atomic
@login_required(login_url = settings.LOGIN_URL)
def hapus_soal(request):
	success = True
	if request.POST:
		id_soal = request.POST.get('id_soal')
		soal_obj = get_object_or_404(Soal, pk=id_soal)	
		try:
			with transaction.atomic():
				soal_obj.delete()
				messages.add_message(request, messages.SUCCESS, '<strong>Sukses! </strong> Data soal berhasil dihapus')
				return redirect('list_soal')
		except DatabaseError:
			messages.add_message(request, messages.WARNING, '<strong>Error! </strong> Terjadi masalah saat meyimpan data, silakan hub IT')
			success = False

@csrf_protect
@transaction.atomic
@login_required(login_url = settings.LOGIN_URL)
def tambah_soal(request):
	data = {}
	data['app_title'] = 'ADMIN | TAMBAH SOAL'	
	data['soal_form'] = SoalForm()
	JawabanFormSet = inlineformset_factory(Soal, Jawaban, fields=('text_jawaban', 'gambar','jawaban_benar', ), extra=5, can_delete=False)
	data['jawaban_form'] = JawabanFormSet
	success = True
	if request.POST:
		f = SoalForm(request.POST, request.FILES)
		if f.is_valid():
			soal_form = f.save(commit = False)
			soal_form.user_insert = request.user
			soal_form.save()					
			try:
				with transaction.atomic():
					if soal_form.tipe_soal == 1:						
						j = JawabanFormSet(request.POST, request.FILES, instance = soal_form)
						j.save()
					else:
						jawaban = ['A','B','C','D','E']
						jawab = request.POST.get('jawaban')
						for j in jawaban:
							if jawab == j.lower():
								benar = True
							else:
								benar = False
							j_formset = Jawaban(soal = soal_form, text_jawaban = j, jawaban_benar = benar)
							j_formset.save()
					#get id soal terakhir
					id_soal = Soal.objects.latest('id')
					messages.add_message(request, messages.SUCCESS, '<strong>Sukses! </strong> Data soal berhasil ditambahkan')
					return redirect('view_soal', id_soal.id)					
			except DatabaseError:
				success = False
				messages.add_message(request, messages.WARNING, '<strong>Error! </strong> Terjadi masalah saat meyimpan data, silakan hub IT')		
	return render(request, 'soal/tambah_soal.html', data)

@csrf_protect
@transaction.atomic
@login_required(login_url = settings.LOGIN_URL)
def tambah_kategori_soal(request):
	data = {}
	data['app_title'] = 'Tambah Kategori Soal'
	data['kategori_form'] = KategoriSoalForm()
	data['kategori_soal'] = Kategori_soal.objects.all()
	if request.method == 'POST':
		k_s_form = KategoriSoalForm(request.POST)
		if k_s_form.is_valid():
			try:
				with transaction.atomic():
					k_s_form.save()
					messages.add_message(request, messages.SUCCESS, '<strong>Sukses! </strong> Data kategori soal berhasil ditambahkan')
					return HttpResponseRedirect(reverse('tambah_kategori_soal'))
			except DatabaseError:
				messages.add_message(request, messages.WARNING, '<strong>Error! </strong> Terjadi masalah saat meyimpan data, silakan hub IT')						
	return render(request, 'soal/tambah_kategori_soal.html', data)

@csrf_protect
@transaction.atomic
@login_required(login_url = settings.LOGIN_URL)
def ubah_kategori_soal(request, id_kategori):
	data = {}
	kategori = get_object_or_404(Kategori_soal, pk = id_kategori)	
	data['kategori_form'] = KategoriSoalForm(instance = kategori)
	if request.method == 'POST':
		k_s_form = KategoriSoalForm(request.POST, instance = kategori)
		if k_s_form.is_valid():
			try:
				with transaction.atomic():
					k_s_form.save()
					messages.add_message(request, messages.SUCCESS, '<strong>Sukses! </strong> Data kategori soal berhasil diubah')
					return HttpResponseRedirect(reverse('tambah_kategori_soal'))
			except DatabaseError:
				messages.add_message(request, messages.WARNING, '<strong>Error! </strong> Terjadi masalah saat meyimpan data, silakan hub IT')
	return render(request, 'soal/ubah_kategori_soal.html', data)

@csrf_protect
@transaction.atomic
@login_required(login_url = settings.LOGIN_URL)
def hapus_kategori_soal(request):
	if request.method == "POST":
		id_kategori = request.POST.get('id_kategori')
		print(id_kategori)
		kategori = get_object_or_404(Kategori_soal, pk = id_kategori)
		kategori.delete()
	return HttpResponseRedirect(reverse('tambah_kategori_soal'))

@csrf_protect
@transaction.atomic
@login_required(login_url = settings.LOGIN_URL)
def peserta_ujian(request):
	data = {}
	peserta = Peserta.objects.all()
	data['app_title'] = 'Peserta Ujian'
	data['mata_kuliah'] = Mata_kuliah.objects.all()
	if request.method == 'POST':
		mata_kuliah = request.POST.get('mata_kuliah')
		peserta = Peserta.objects.filter(mata_kuliah = mata_kuliah)
				
	data['peserta'] = peserta
	return render(request, 'soal/peserta_ujian.html', data)

@csrf_protect
@transaction.atomic
@login_required(login_url = settings.LOGIN_URL)
def upload_peserta(request):
	data = {}
	excel = {}
	excel_list = []
	data['app_title'] = 'Upload Peserta Ujian'	
	if request.method == 'POST':		
		peserta_form = UploadPesertaForm(request.POST, request.FILES)
		mata_kuliah = get_object_or_404(Mata_kuliah, pk=request.POST.get('mata_kuliah')) 		
		if peserta_form.is_valid():
			wb = load_workbook(request.FILES['file_csv'], data_only=True)
			sheet = wb.active			
			max_row = str(sheet.max_row)
			cells = sheet['A2':'C'+max_row]
			for nim, nama, prodi in cells:
				if nim.value is not None or nama.value is not None or prodi.value is not None:
					nim = nim.value
					nama = nama.value
					prodi = prodi.value
					user, created = User.objects.get_or_create(email='jlennon@beatles.com', username = nim)
					if created:
						user.set_password(nim)
					user.save()
					user_id = get_object_or_404(User, username = nim)

					try:
						with transaction.atomic():
							p = Peserta.objects.get_or_create(nim = nim, nama = nama, mata_kuliah = mata_kuliah, program_studi = prodi, uniid = user_id)
					except DatabaseError:
						messages.add_message(request, messages.WARNING, '<strong>Error! </strong> Terjadi masalah saat meyimpan data, silakan hub IT')						
				else:
					continue
				excel_list.append({'nim':nim, 'nama': nama, 'prodi': prodi, 'mata_kuliah': mata_kuliah})			
	else:
		peserta_form = UploadPesertaForm()	
			
	data['peserta_form'] = peserta_form
	data['excel'] = excel_list
	return render(request, 'soal/upload_peserta.html', data)

@csrf_protect
@transaction.atomic
@login_required(login_url = settings.LOGIN_URL)
def hapus_peserta(request):
	data = {}
	if request.method == "POST":
		hapus_id = request.POST.getlist('_selected_action')		
		if not hapus_id:			
			return HttpResponseRedirect(reverse('peserta_ujian'))
		else:
			hapus_peserta = Peserta.objects.filter(pk__in = hapus_id)			
			if request.POST.get('_confirm_delete') == '1':
				hapus_peserta.delete()
				return HttpResponseRedirect(reverse('peserta_ujian'))
	data['hapus_peserta'] = hapus_peserta
	return render(request, 'soal/konfirmasi_hapus_peserta.html', data)

@csrf_protect
@transaction.atomic
@login_required(login_url = settings.LOGIN_URL)
def ubah_peserta(request, id_peserta):
	data = {}
	peserta = get_object_or_404(Peserta, pk=id_peserta)
	peserta_form = PesertaForm(instance = peserta)
	if request.method == 'POST':
		p_form = PesertaForm(request.POST, instance = peserta)
		try:
			with transaction.atomic():
				p_form.save()
				messages.add_message(request, messages.SUCCESS, '<strong>Sukses! </strong> Data peserta ujian berhasil diubah')
				return HttpResponseRedirect(reverse('peserta_ujian'))
		except DatabaseError:
			messages.add_message(request, messages.WARNING, '<strong>Error! </strong> Terjadi masalah saat meyimpan data, silakan hub IT')
			
	data['peserta_form'] = peserta_form
	return render(request, 'soal/ubah_peserta.html', data)